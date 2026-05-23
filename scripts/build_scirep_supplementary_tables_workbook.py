from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TABLES = [
    ("S1", "Parameterized stressor levels", "paper/scientific_reports/tables/stressor_parameter_table.tex", None),
    ("S2", "Benchmark comparison", "paper/scientific_reports/tables/benchmark_comparison_table.tex", None),
    ("S3", "Claim-support matrix", "paper/scientific_reports/tables/claim_support_matrix_full.tex", None),
    ("S4", "Oracle diagnostic success", "paper/scientific_reports/tables/oracle_diagnostic_success_summary.tex", "paper/scientific_reports/tables/oracle_diagnostic_success_summary.csv"),
    ("S5", "Threshold sensitivity", "paper/scientific_reports/tables/threshold_sensitivity.tex", "scientific_reports_revision_20260521/threshold_sensitivity.csv"),
    ("S6", "Detector-bridge summary", "paper/scientific_reports/tables/open_vocab_detector_transfer_summary.tex", "paper/scientific_reports/tables/open_vocab_detector_transfer_summary.csv"),
    ("S7", "YCB/clutter query ablation", "paper/scientific_reports/tables/query_ablation_summary.tex", "paper/scientific_reports/tables/query_ablation_summary.csv"),
    ("S8", "Object-name resolution audit", "paper/scientific_reports/tables/target_name_probe_summary.tex", "paper/scientific_reports/tables/target_name_probe_summary.csv"),
    ("S9", "Paired effect summaries", "paper/scientific_reports/tables/paired_effects_scirep.tex", "paper/scientific_reports/tables/paired_effects_scirep.csv"),
    ("S10", "Low-IoU valid-3D summary", "paper/scientific_reports/tables/low_iou_valid_3d_summary.tex", "paper/scientific_reports/tables/low_iou_valid_3d_summary.csv"),
    ("S11", "External RGB-D validation summary", "paper/scientific_reports/tables/external_rgbd_validation_summary.tex", "paper/scientific_reports/tables/external_rgbd_validation_summary.csv"),
    ("S12", "External RGB-D object stratification", "paper/scientific_reports/tables/external_rgbd_by_object.tex", "paper/scientific_reports/tables/external_rgbd_by_object.csv"),
    ("S13", "External RGB-D scene stratification", "paper/scientific_reports/tables/external_rgbd_by_scene.tex", "paper/scientific_reports/tables/external_rgbd_by_scene.csv"),
    ("S14", "External RGB-D detector-threshold sensitivity", "paper/scientific_reports/tables/external_rgbd_detector_threshold_sensitivity.tex", "paper/scientific_reports/tables/external_rgbd_detector_threshold_sensitivity.csv"),
]


def _extract_caption(tex: str) -> str:
    marker = r"\caption{"
    start = tex.find(marker)
    if start < 0:
        return ""
    i = start + len(marker)
    depth = 1
    chars: list[str] = []
    while i < len(tex) and depth:
        ch = tex[i]
        if ch == "{":
            depth += 1
            chars.append(ch)
        elif ch == "}":
            depth -= 1
            if depth:
                chars.append(ch)
        else:
            chars.append(ch)
        i += 1
    return _latex_to_text("".join(chars))


def _latex_to_text(text: str) -> str:
    replacements = {
        r"\%": "%",
        r"\&": "&",
        r"\_": "_",
        r"~": " ",
        r"--": "-",
        r"``": '"',
        r"''": '"',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"\\(?:texttt|textbf|emph)\{([^{}]*)\}", r"\1", text)
    text = re.sub(r"\\cite\{([^{}]*)\}", r"\1", text)
    text = re.sub(r"\\(?:ref|url|href)\{([^{}]*)\}(?:\{([^{}]*)\})?", lambda m: m.group(2) or m.group(1), text)
    text = re.sub(r"\\[a-zA-Z]+\*?(?:\[[^\]]*\])?(?:\{[^{}]*\})?", "", text)
    text = text.replace("{", "").replace("}", "")
    return re.sub(r"\s+", " ", text).strip()


def _split_latex_row(row: str) -> list[str]:
    cells: list[str] = []
    current: list[str] = []
    escaped = False
    for ch in row:
        if escaped:
            current.append("\\" + ch)
            escaped = False
        elif ch == "\\":
            escaped = True
        elif ch == "&":
            cells.append(_latex_to_text("".join(current)))
            current = []
        else:
            current.append(ch)
    if escaped:
        current.append("\\")
    cells.append(_latex_to_text("".join(current)))
    return cells


def _rows_from_tex(tex_path: Path) -> list[list[str]]:
    tex = tex_path.read_text(encoding="utf-8")
    match = re.search(r"\\begin\{tabular\}.*?\\end\{tabular\}", tex, re.S)
    if not match:
        return []
    body = match.group(0)
    body = re.sub(r"\\begin\{tabular\}\{[^}]*\}", "", body, flags=re.S)
    body = body.replace(r"\end{tabular}", "")
    body = re.sub(r"\\(?:toprule|midrule|bottomrule|hline)", "", body)
    body = re.sub(r"\\setlength\{[^{}]*\}\{[^{}]*\}", "", body)
    rows: list[list[str]] = []
    for raw in re.split(r"\\\\", body):
        raw = raw.strip()
        if not raw or raw.startswith(r"\begin") or raw.startswith(r"\end"):
            continue
        if "&" not in raw:
            continue
        rows.append(_split_latex_row(raw))
    return rows


def _rows_from_csv(csv_path: Path) -> list[list[object]]:
    df = pd.read_csv(csv_path)
    return [list(df.columns)] + df.fillna("").astype(object).values.tolist()


def _write_sheet(wb: Workbook, sheet_name: str, title: str, caption: str, rows: list[list[object]], source: str) -> None:
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = f"Supplementary Table {sheet_name}: {title}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = caption
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A3"] = f"Source: {source}"
    ws["A3"].font = Font(italic=True, color="666666")
    start_row = 5
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for r_idx, row in enumerate(rows, start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if r_idx == start_row:
                cell.font = Font(bold=True)
                cell.fill = header_fill
    ws.freeze_panes = f"A{start_row + 1}"
    max_cols = max((len(row) for row in rows), default=1)
    for c_idx in range(1, max_cols + 1):
        values = [str(ws.cell(row=r, column=c_idx).value or "") for r in range(1, min(ws.max_row, 80) + 1)]
        width = min(max(max((len(v) for v in values), default=10) + 2, 12), 42)
        ws.column_dimensions[get_column_letter(c_idx)].width = width


def build(output_xlsx: Path, manifest_csv: Path, manifest_md: Path) -> None:
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "Manifest"
    ws_meta.append(["Supplementary table", "Title", "Caption", "Source table", "Source data"])
    for cell in ws_meta[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    manifest_rows: list[dict[str, str]] = []
    for number, title, tex_name, csv_name in TABLES:
        tex_path = Path(tex_name)
        csv_path = Path(csv_name) if csv_name else None
        caption = _extract_caption(tex_path.read_text(encoding="utf-8")) if tex_path.exists() else ""
        if csv_path and csv_path.exists():
            rows = _rows_from_csv(csv_path)
            source = str(csv_path)
        else:
            rows = _rows_from_tex(tex_path)
            source = str(tex_path)
        _write_sheet(wb, number, title, caption, rows, source)
        manifest_row = {
            "supplementary_table": number,
            "title": title,
            "caption": caption,
            "source_table": tex_name,
            "source_data": csv_name or tex_name,
        }
        manifest_rows.append(manifest_row)
        ws_meta.append([number, title, caption, tex_name, csv_name or tex_name])
    for col in range(1, 6):
        ws_meta.column_dimensions[get_column_letter(col)].width = [18, 34, 80, 54, 54][col - 1]
    for row in ws_meta.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(output_xlsx)

    manifest_csv.parent.mkdir(parents=True, exist_ok=True)
    with manifest_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["supplementary_table", "title", "caption", "source_table", "source_data"])
        writer.writeheader()
        writer.writerows(manifest_rows)

    lines = ["# Supplementary Tables Manifest", ""]
    lines.append("This manifest maps every supplementary table cited in the Scientific Reports manuscript to its source table and source data.")
    lines.append("")
    lines.append("| Table | Title | Source data |")
    lines.append("| --- | --- | --- |")
    for row in manifest_rows:
        lines.append(f"| {row['supplementary_table']} | {row['title']} | `{row['source_data']}` |")
    manifest_md.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-xlsx", default="paper/scientific_reports/supplementary_tables_S1_S14.xlsx")
    parser.add_argument("--manifest-csv", default="paper/scientific_reports/supplementary_tables_manifest.csv")
    parser.add_argument("--manifest-md", default="paper/scientific_reports/supplementary_tables_manifest.md")
    args = parser.parse_args()
    build(Path(args.output_xlsx), Path(args.manifest_csv), Path(args.manifest_md))
    print(f"Wrote {args.output_xlsx}")
    print(f"Wrote {args.manifest_csv}")
    print(f"Wrote {args.manifest_md}")


if __name__ == "__main__":
    main()
